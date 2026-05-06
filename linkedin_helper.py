import anthropic
import os
import json
import re
import openpyxl
from dotenv import load_dotenv

load_dotenv()

EXCEL_FILE      = "job_search_tracker.xlsx"
OUTREACH_FILE   = "outreach_prep.json"
client          = anthropic.Anthropic(api_key=os.getenv("ANTHROPIC_API_KEY"))


# -------------------------------------------------------------------
# LOAD TODAY'S JOBS FROM OUTREACH PREP
# -------------------------------------------------------------------
def load_outreach_jobs():
    if not os.path.exists(OUTREACH_FILE):
        print("  ⚠️  outreach_prep.json not found!")
        print("       Run python main.py first to prepare today's jobs.")
        return []

    with open(OUTREACH_FILE, "r") as f:
        return json.load(f)


# -------------------------------------------------------------------
# FETCH PROFILE INFO VIA WEB SEARCH
# Claude uses web search to find public info about the person
# -------------------------------------------------------------------
def fetch_profile_info(profile_url, job_context):
    """
    Uses Claude with web search to find info about
    the person from their LinkedIn URL.
    Returns structured profile data.
    """
    print("\n  🔍 Searching for profile info...")

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1000,
        tools=[{
            "type": "web_search_20250305",
            "name": "web_search"
        }],
        messages=[{
            "role": "user",
            "content": f"""Search for information about this LinkedIn profile:
{profile_url}

Find and return ONLY this information in JSON format:
{{
  "name": "their full name",
  "first_name": "first name only",
  "title": "their current job title",
  "company": "their current company",
  "location": "their location",
  "education": "their university if found",
  "recent_activity": "any recent post or article topic if found",
  "years_at_company": "how long at current company if found",
  "background_hook": "one specific interesting detail that could be used to personalize a message"
}}

Return ONLY raw JSON, no markdown, no backticks."""
        }]
    )

    # Extract text from response
    full_text = ""
    for block in response.content:
        if hasattr(block, "text"):
            full_text += block.text

    # Parse JSON
    try:
        raw = re.sub(r"```json|```", "", full_text).strip()
        return json.loads(raw)
    except:
        # Fallback if parsing fails
        print("  ⚠️  Could not auto-fetch profile. Please paste info manually.")
        return None


# -------------------------------------------------------------------
# MANUAL PROFILE INPUT FALLBACK
# If web search doesn't find enough info
# -------------------------------------------------------------------
def get_profile_manually():
    """Ask user to paste key profile details"""
    print("\n  📋 Please paste their profile details:")
    print("     (Name, Title, Company, anything notable)")
    print("     Press Enter twice when done:\n")

    lines = []
    while True:
        line = input("     ")
        if line == "" and lines and lines[-1] == "":
            break
        lines.append(line)

    raw_text = "\n".join(lines)

    # Use Claude to structure what was pasted
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=500,
        messages=[{
            "role": "user",
            "content": f"""Extract info from this LinkedIn profile text and return ONLY JSON:
{{
  "name": "",
  "first_name": "",
  "title": "",
  "company": "",
  "location": "",
  "education": "",
  "recent_activity": "",
  "background_hook": "one interesting detail for personalization"
}}

Profile text:
{raw_text}

Return ONLY raw JSON, no markdown."""
        }]
    )

    try:
        raw = re.sub(r"```json|```", "",
                     response.content[0].text).strip()
        return json.loads(raw)
    except:
        return {"name": "there", "first_name": "there",
                "title": "", "company": "", "location": "",
                "education": "", "recent_activity": "",
                "background_hook": ""}


# -------------------------------------------------------------------
# GENERATE PERSONALIZED DMs
# Claude writes connection note + follow-up based on profile
# -------------------------------------------------------------------
def generate_dms(profile, job, person_type):
    """
    Generates hyper-personalized:
    1. Connection note (max 280 chars)
    2. Follow-up DM (after they accept)
    """

    # Determine tone based on person type
    tone_guide = {
        "recruiter": "Professional, direct, focused on being considered for the role. Mention you applied.",
        "manager":   "Respectful, curious about the team, highlight relevant experience briefly.",
        "peer":      "Friendly, collegial, express genuine interest in their experience at the company."
    }
    tone = tone_guide.get(person_type, tone_guide["peer"])

    print(f"\n  ✍️  Generating personalized DMs for {profile.get('first_name')}...")

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=1000,
        messages=[{
            "role": "user",
            "content": f"""You are an expert at writing personalized LinkedIn messages
that get responses. Write two messages for this situation.

PERSON DETAILS:
- Name: {profile.get('name')}
- First name: {profile.get('first_name')}
- Title: {profile.get('title')}
- Company: {profile.get('company')}
- Location: {profile.get('location')}
- Education: {profile.get('education')}
- Recent activity: {profile.get('recent_activity')}
- Interesting hook: {profile.get('background_hook')}

JOB APPLIED FOR:
- Title: {job.get('title')}
- Company: {job.get('company')}
- ATS Score: {job.get('score')}/100

SENDER PROFILE (Hitankshi Jain):
- 4 years Business Analyst experience
- MS Business Analytics, Drexel University, Philadelphia
- Skills: SQL, Python, PowerBI, Tableau, R
- Currently at Neuralix.ai as Business Analyst
- Previously: Tech Impact (BA Intern), TCS (BA)

TONE: {tone}

RULES — CRITICAL:
1. Connection note MUST be under 280 characters (count carefully)
2. Use their first name naturally
3. Reference ONE specific detail from their profile as a hook
4. Sound completely human — not like a template
5. Never say "I came across your profile"
6. Never use hollow phrases like "I hope this message finds you well"
7. Connection note must mention the specific role at the company
8. Follow-up DM should be sent after they accept — more detailed
9. Follow-up should be 3-4 sentences max
10. Both must feel genuine and specific to THIS person

Return ONLY raw JSON, no markdown:
{{
  "connection_note": "the connection note text here",
  "connection_note_chars": 0,
  "follow_up_dm": "the follow up message here",
  "personalization_score": 0,
  "personalization_reason": "why this is personalized"
}}"""
        }]
    )

    try:
        raw = re.sub(r"```json|```",
                     "", response.content[0].text).strip()
        result = json.loads(raw)

        # Double check character count
        note  = result.get("connection_note", "")
        count = len(note)
        result["connection_note_chars"] = count

        # Warn if over limit
        if count > 300:
            print(f"  ⚠️  Note is {count} chars — over LinkedIn limit!")
        
        return result

    except Exception as e:
        print(f"  ❌ Error generating DMs: {e}")
        return None


# -------------------------------------------------------------------
# SAVE TO EXCEL
# Updates the job row with person info + DMs
# -------------------------------------------------------------------
def save_to_excel(job, profile, dms, profile_url, person_number):
    """
    Updates Excel tracker with LinkedIn outreach info.
    Adds to the correct job row.
    """
    if not os.path.exists(EXCEL_FILE):
        print("  ⚠️  Excel file not found!")
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["Job Applications"]

    # Find the column positions — add if not exist
    headers = [cell.value for cell in ws[1]]

    # Define new columns we need
    new_cols = [
        f"Person {person_number} Name",
        f"Person {person_number} Title",
        f"Person {person_number} Profile URL",
        f"Person {person_number} Connection Note",
        f"Person {person_number} Chars",
        f"Person {person_number} Follow-up DM",
        f"Person {person_number} Sent?"
    ]

    # Add columns if they don't exist
    for col_name in new_cols:
        if col_name not in headers:
            ws.cell(row=1, column=len(headers) + 1,
                    value=col_name)
            headers.append(col_name)

    # Find the job row by company + title match
    target_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        cell_title   = row[2].value  # column C = Job Title
        cell_company = row[3].value  # column D = Company
        if (cell_title == job.get("title") and
                cell_company == job.get("company")):
            target_row = row[0].row
            break

    if not target_row:
        print(f"  ⚠️  Could not find {job.get('title')} @ "
              f"{job.get('company')} in Excel")
        return

    # Write data to found row
    col_map = {h: i+1 for i, h in enumerate(headers)}

    updates = {
        f"Person {person_number} Name":
            profile.get("name", ""),
        f"Person {person_number} Title":
            profile.get("title", ""),
        f"Person {person_number} Profile URL":
            profile_url,
        f"Person {person_number} Connection Note":
            dms.get("connection_note", ""),
        f"Person {person_number} Chars":
            dms.get("connection_note_chars", 0),
        f"Person {person_number} Follow-up DM":
            dms.get("follow_up_dm", ""),
        f"Person {person_number} Sent?":
            "No"
    }

    for col_name, value in updates.items():
        if col_name in col_map:
            ws.cell(row=target_row,
                    column=col_map[col_name],
                    value=value)

    wb.save(EXCEL_FILE)
    print(f"  📊 Excel updated with {profile.get('name')}'s info!")


# -------------------------------------------------------------------
# MAIN INTERACTIVE TOOL
# -------------------------------------------------------------------
def run_helper():
    print("\n" + "="*55)
    print("  🤝 LINKEDIN OUTREACH HELPER")
    print("  Paste a profile URL → get personalized DM")
    print("="*55)

    # Load today's jobs
    jobs = load_outreach_jobs()
    if not jobs:
        return

    # Show available jobs
    print("\n  📋 Today's jobs ready for outreach:\n")
    for i, job in enumerate(jobs):
        print(f"  [{i+1}] {job['title']} @ {job['company']} "
              f"— Score: {job['score']}/100")

    print()

    while True:
        # Pick a job
        try:
            choice = input("  Select job number (or 'q' to quit): ").strip()
            if choice.lower() == 'q':
                print("\n  👋 Done! Check Excel for all saved DMs.")
                break

            job_idx = int(choice) - 1
            if job_idx < 0 or job_idx >= len(jobs):
                print("  ⚠️  Invalid number. Try again.")
                continue

            job = jobs[job_idx]
            print(f"\n  ✅ Selected: {job['title']} @ {job['company']}")

            # Show search URLs
            print("\n  🔗 Open these LinkedIn searches to find people:")
            urls = job.get("search_urls", {})
            for ptype, info in urls.items():
                print(f"     {info['label']}:")
                print(f"     {info['url']}\n")

        except ValueError:
            print("  ⚠️  Please enter a number.")
            continue

        # Process people for this job
        person_number = 1
        while person_number <= 3:
            print(f"\n  --- Person {person_number} of 3 ---")
            print("  Find someone on LinkedIn using the links above.")
            print("  Then paste their profile URL below.\n")

            profile_url = input(
                "  Paste LinkedIn profile URL "
                "(or 'skip' to skip, 'done' to finish job): "
            ).strip()

            if profile_url.lower() == 'done':
                break
            if profile_url.lower() == 'skip':
                person_number += 1
                continue
            if not profile_url:
                continue

            # Validate it's a LinkedIn URL
            if "linkedin.com/in/" not in profile_url:
                print("  ⚠️  Please paste a valid LinkedIn profile URL")
                print("       e.g. https://www.linkedin.com/in/sarah-johnson")
                continue

            # Determine person type
            print("\n  What type of person is this?")
            print("  [1] Recruiter / Talent Acquisition")
            print("  [2] Senior / Hiring Manager")
            print("  [3] Same title / Peer")
            type_choice = input("  Enter 1, 2 or 3: ").strip()

            type_map = {"1": "recruiter", "2": "manager", "3": "peer"}
            person_type = type_map.get(type_choice, "peer")

            # Try to fetch profile via web search
            profile = fetch_profile_info(profile_url, job)

            # Fallback to manual if web search fails
            if not profile or not profile.get("name"):
                print("  ℹ️  Auto-fetch didn't find enough info.")
                profile = get_profile_manually()

            # Show what we found
            print("\n  👤 Profile found:")
            print(f"     Name:     {profile.get('name')}")
            print(f"     Title:    {profile.get('title')}")
            print(f"     Company:  {profile.get('company')}")
            print(f"     Hook:     {profile.get('background_hook')}")

            # Generate DMs
            dms = generate_dms(profile, job, person_type)

            if dms:
                print(f"\n  {'='*50}")
                print(f"  ✉️  CONNECTION NOTE "
                      f"({dms.get('connection_note_chars')}/300 chars):")
                print(f"  {'='*50}")
                print(f"\n  {dms.get('connection_note')}\n")

                print(f"  {'='*50}")
                print("  📩 FOLLOW-UP DM (send after they accept):")
                print(f"  {'='*50}")
                print(f"\n  {dms.get('follow_up_dm')}\n")

                print(f"  💡 Personalization score: "
                      f"{dms.get('personalization_score')}/100")
                print(f"     {dms.get('personalization_reason')}")

                # Save to Excel
                save_choice = input(
                    "\n  Save to Excel? (y/n): "
                ).strip().lower()

                if save_choice == 'y':
                    save_to_excel(
                        job, profile, dms,
                        profile_url, person_number
                    )

            person_number += 1

            # Ask if want another person for same job
            if person_number <= 3:
                another = input(
                    "\n  Add another person for this job? (y/n): "
                ).strip().lower()
                if another != 'y':
                    break

        # Ask if want another job
        another_job = input(
            "\n  Move to another job? (y/n): "
        ).strip().lower()
        if another_job != 'y':
            print("\n  👋 Done! Check Excel for all saved DMs.")
            print(f"  📊 Open: {EXCEL_FILE}")
            break


if __name__ == "__main__":
    run_helper()
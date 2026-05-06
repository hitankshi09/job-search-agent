import json
import os
import time
from datetime import datetime
from playwright.sync_api import sync_playwright
from dotenv import load_dotenv

load_dotenv()

EXCEL_FILE    = "job_search_tracker.xlsx"
PROFILE_FILE  = "candidate_profile.json"


# -------------------------------------------------------------------
# LOAD CANDIDATE PROFILE
# -------------------------------------------------------------------
def load_candidate():
    if os.path.exists(PROFILE_FILE):
        with open(PROFILE_FILE, "r") as f:
            return json.load(f)
    return {}


def get_resume_data(result):
    return result.get("app_data", {})


# -------------------------------------------------------------------
# SHOW REFERENCE PANEL
# -------------------------------------------------------------------
def show_reference_panel(data, candidate):
    print(f"\n  {'='*55}")
    print(f"  📋 COPY-PASTE REFERENCE")
    print(f"  {'='*55}")
    print(f"  Name:      {data.get('full_name')}")
    print(f"  Email:     {data.get('email')}")
    print(f"  Phone:     {data.get('phone')}")
    print(f"  Location:  {data.get('location')}")
    print(f"  LinkedIn:  {data.get('linkedin_url')}")
    print(f"  GitHub:    {data.get('github_url')}")
    print(f"  Visa:      {candidate.get('visa_sponsorship')}")
    print(f"  Relocate:  {candidate.get('open_to_relocation')}")

    print(f"\n  EDUCATION:")
    for i, edu in enumerate(data.get("education", []), 1):
        print(f"  [{i}] {edu.get('university')} | "
              f"{edu.get('degree')} {edu.get('field')}")
        print(f"      {edu.get('start_date')} → {edu.get('end_date')}")

    print(f"\n  WORK EXPERIENCE:")
    for i, exp in enumerate(data.get("work_experience", []), 1):
        print(f"  [{i}] {exp.get('company')} | {exp.get('title')}")
        print(f"      {exp.get('start_date')} → {exp.get('end_date')}")

    bullets = data.get("tailored_bullets", [])
    if bullets:
        print(f"\n  TAILORED BULLETS:")
        for i, b in enumerate(bullets[:4], 1):
            print(f"  [{i}] {b[:90]}")

    print(f"\n  SKILLS: {', '.join(data.get('skills', []))}")
    print(f"  {'='*55}\n")


# -------------------------------------------------------------------
# CORE HELPERS
# -------------------------------------------------------------------
def safe_fill(page, selector, value, timeout=2000):
    """Safely fill a text input"""
    if not value:
        return False
    try:
        el = page.wait_for_selector(selector, timeout=timeout)
        if el and el.is_visible():
            el.click()
            el.fill("")
            el.fill(str(value))
            return True
    except:
        pass
    return False


def safe_select(page, selector, value, timeout=2000):
    """
    Safely select from a standard HTML <select> element.
    Tries exact label match first, then partial match.
    """
    if not value:
        return False
    try:
        el = page.wait_for_selector(selector, timeout=timeout)
        if not el or not el.is_visible():
            return False

        # Try exact label match
        try:
            page.select_option(selector, label=value)
            return True
        except:
            pass

        # Try partial label match — get all options and find closest
        options = page.query_selector_all(f"{selector} option")
        for opt in options:
            opt_text = opt.inner_text().strip()
            if value.lower() in opt_text.lower():
                opt_val = opt.get_attribute("value")
                if opt_val:
                    page.select_option(selector, value=opt_val)
                    return True
    except:
        pass
    return False


def safe_upload(page, selector, file_path, timeout=3000):
    """Safely upload a file to a standard file input"""
    if not file_path or not os.path.exists(file_path):
        return False
    try:
        page.wait_for_selector(selector, timeout=timeout)
        page.set_input_files(selector, file_path)
        return True
    except:
        pass
    return False


def try_fill_any(page, selectors, value, timeout=2000):
    """Try multiple selectors — return True on first success"""
    for sel in selectors:
        if safe_fill(page, sel, value, timeout):
            return True
    return False


def try_select_any(page, selectors, value, timeout=2000):
    """Try multiple select selectors — return True on first success"""
    for sel in selectors:
        if safe_select(page, sel, value, timeout):
            return True
    return False


# -------------------------------------------------------------------
# DETECT PORTAL
# -------------------------------------------------------------------
def detect_portal(url):
    u = url.lower()
    if "greenhouse.io" in u or "boards.greenhouse" in u:
        return "greenhouse"
    elif "myworkdayjobs.com" in u or "workday.com" in u:
        return "workday"
    elif "jobs.lever.co" in u or "lever.co" in u:
        return "lever"
    elif "icims.com" in u:
        return "icims"
    elif "taleo.net" in u:
        return "taleo"
    elif "linkedin.com/jobs" in u:
        return "linkedin"
    else:
        return "unknown"


# -------------------------------------------------------------------
# GREENHOUSE FILLER
# Confirmed selectors from Greenhouse API docs + ClawHub research:
# input#first_name, input#last_name, input#email, input#phone
# Resume: expect_file_chooser with Attach button
# EEO: select#gender, select#race, select#veteran_status,
#       select#disability_status — standard HTML selects
# -------------------------------------------------------------------
def fill_greenhouse(page, data, candidate, resume_path):
    print("  📋 Filling Greenhouse form...")
    filled = []

    # Wait for form to load
    page.wait_for_load_state("domcontentloaded")
    time.sleep(1)

    # --- Basic info ---
    if safe_fill(page, "input#first_name",
                 data.get("first_name", "")):
        filled.append("First name")

    if safe_fill(page, "input#last_name",
                 data.get("last_name", "")):
        filled.append("Last name")

    if safe_fill(page, "input#email",
                 data.get("email", "")):
        filled.append("Email")

    if safe_fill(page, "input#phone",
                 data.get("phone", "")):
        filled.append("Phone")

    # Location — Greenhouse uses autocomplete
    if safe_fill(page, "input#job_application_location",
                 data.get("location", "")):
        time.sleep(0.5)
        page.keyboard.press("Escape")  # close autocomplete
        filled.append("Location")

    # LinkedIn URL
    if try_fill_any(page, [
        "input[name*='linkedin' i]",
        "input[id*='linkedin' i]",
        "input[placeholder*='linkedin' i]",
        "input[aria-label*='linkedin' i]"
    ], data.get("linkedin_url", "")):
        filled.append("LinkedIn")

    # GitHub URL
    if try_fill_any(page, [
        "input[name*='github' i]",
        "input[id*='github' i]",
        "input[placeholder*='github' i]"
    ], data.get("github_url", "")):
        filled.append("GitHub")

    # --- Resume upload ---
    # First try standard file input
    resume_uploaded = False
    if safe_upload(page, "input[type='file']", resume_path):
        filled.append("✅ Resume")
        resume_uploaded = True

    # If not found — use Attach button with file chooser
    if not resume_uploaded:
        try:
            with page.expect_file_chooser(timeout=4000) as fc:
                page.click("text=Attach")
            fc.value.set_files(resume_path)
            filled.append("✅ Resume (Attach)")
            resume_uploaded = True
            time.sleep(1.5)
        except:
            print(f"  ⚠️  Upload manually: {os.path.basename(resume_path)}")

    # --- Scroll down to reach EEO + other fields ---
    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    time.sleep(1)

    # --- EEO fields ---
    # Research confirmed these are standard HTML selects
    # with IDs: gender, race, veteran_status, disability_status
    eeo_filled = fill_greenhouse_eeo(page, candidate)
    filled.extend(eeo_filled)

    # --- Custom questions (visa, relocation etc) ---
    custom_filled = fill_greenhouse_custom(page, candidate)
    filled.extend(custom_filled)

    # Scroll back to top
    page.evaluate("window.scrollTo(0, 0)")

    print(f"  ✅ Greenhouse filled: {', '.join(filled)}")
    return filled


def fill_greenhouse_eeo(page, candidate):
    """
    Greenhouse EEO uses React Select components.
    Correct pattern: click → type → press Enter
    NOT standard select_option()
    """
    filled = []

    def react_select(label_text, value):
        """
        Fill a React Select dropdown by:
        1. Finding the container by label text
        2. Clicking it to open
        3. Typing the value to filter
        4. Pressing Enter to select
        """
        if not value:
            return False
        try:
            # Find the select container near the label
            # Greenhouse React Select structure:
            # <label>Gender</label>
            # <div class="select__control"> ← click this
            #   <input class="select__input"> ← type here
            # <div class="select__menu">
            #   <div class="select__option"> ← Enter selects

            # Method 1: Find by label then navigate to input
            result = page.evaluate(f"""
                () => {{
                    // Find label with matching text
                    const labels = document.querySelectorAll('label, div[class*="label"]');
                    for (const label of labels) {{
                        if (label.textContent.trim().toLowerCase()
                                .includes('{label_text.lower()}')) {{
                            // Find the nearest select input
                            const parent = label.closest('div') ||
                                           label.parentElement;
                            if (parent) {{
                                const input = parent.querySelector(
                                    'input[class*="select"], input[id*="react-select"]'
                                );
                                if (input) {{
                                    input.focus();
                                    input.click();
                                    return input.id || 'found';
                                }}
                                // Try sibling
                                const next = label.nextElementSibling;
                                if (next) {{
                                    const inp = next.querySelector('input');
                                    if (inp) {{
                                        inp.focus();
                                        inp.click();
                                        return inp.id || 'found';
                                    }}
                                }}
                            }}
                        }}
                    }}
                    return null;
                }}
            """)

            if result:
                time.sleep(0.3)
                page.keyboard.type(value[:3])  # type first 3 chars to filter
                time.sleep(0.4)
                page.keyboard.press("Enter")
                time.sleep(0.2)
                return True

            # Method 2: Use Playwright locator with label
            try:
                # Find the combobox associated with the label
                combo = page.get_by_role(
                    "combobox",
                    name=label_text,
                    exact=False
                )
                if combo.count() > 0:
                    combo.first.click()
                    time.sleep(0.3)
                    page.keyboard.type(value[:4])
                    time.sleep(0.4)
                    page.keyboard.press("Enter")
                    return True
            except:
                pass

            # Method 3: Find all React Select inputs and try each
            inputs = page.query_selector_all(
                "input[id*='react-select'], "
                "input[class*='select__input']"
            )
            for inp in inputs:
                # Check if this input is near a label with our text
                try:
                    container = inp.evaluate("""
                        el => el.closest('.field, [class*="select-container"],
                        [class*="select__container"], div'
                        )?.previousElementSibling?.textContent || ''
                    """)
                    if label_text.lower() in container.lower():
                        inp.click()
                        time.sleep(0.3)
                        page.keyboard.type(value[:4])
                        time.sleep(0.4)
                        page.keyboard.press("Enter")
                        return True
                except:
                    continue

        except Exception as e:
            pass
        return False

    # Fill each EEO field using React Select pattern
    if react_select("gender", candidate.get("eeo_gender", "Female")):
        filled.append("Gender")
        time.sleep(0.2)

    if react_select("hispanic", candidate.get("eeo_hispanic",
                    "No, not of Hispanic")):
        filled.append("Hispanic/Latino")
        time.sleep(0.2)

    if react_select("veteran", candidate.get("eeo_veteran",
                    "I am not a protected veteran")):
        filled.append("Veteran")
        time.sleep(0.2)

    if react_select("disability", candidate.get("eeo_disability",
                    "I don't wish to answer")):
        filled.append("Disability")
        time.sleep(0.2)

    return filled


def fill_greenhouse_custom(page, candidate):
    """Fill custom Greenhouse questions using React Select pattern"""
    filled = []

    def react_select_or_standard(keyword, value):
        """Try standard select first then React Select"""
        if not value:
            return False

        # Try standard HTML select
        selectors = [
            f"select[id*='{keyword}' i]",
            f"select[name*='{keyword}' i]",
            f"select[aria-label*='{keyword}' i]"
        ]
        for sel in selectors:
            if safe_select(page, sel, value):
                return True

        # Try React Select pattern
        try:
            inputs = page.query_selector_all(
                "input[id*='react-select'], "
                "input[class*='select__input'], "
                f"input[aria-label*='{keyword}' i]"
            )
            for inp in inputs:
                aria = (inp.get_attribute("aria-label") or "").lower()
                if keyword.lower() in aria:
                    inp.click()
                    time.sleep(0.3)
                    page.keyboard.type(value[:4])
                    time.sleep(0.4)
                    page.keyboard.press("Enter")
                    return True
        except:
            pass
        return False

    # Visa sponsorship
    if react_select_or_standard("visa",
                                candidate.get("visa_sponsorship", "Yes")):
        filled.append("Visa")

    if react_select_or_standard("sponsorship",
                                candidate.get("visa_sponsorship", "Yes")):
        if "Visa" not in filled:
            filled.append("Visa2")

    # Relocation
    if react_select_or_standard("relocat",
                                candidate.get("open_to_relocation", "No")):
        filled.append("Relocation")

    # Address
    if try_fill_any(page, [
        "input[id*='address' i]",
        "input[name*='address' i]",
        "input[placeholder*='address' i]",
        "input[placeholder*='working' i]",
        "textarea[placeholder*='address' i]"
    ], candidate.get("address", "")):
        filled.append("Address")

    return filled


# -------------------------------------------------------------------
# WORKDAY FILLER
# Research confirmed: data-automation-id attributes
# Upload resume first → wait 4-5s → fills many fields automatically
# -------------------------------------------------------------------
def fill_workday(page, data, candidate, resume_path):
    print("  📋 Filling Workday form...")
    filled = []

    time.sleep(2)

    # Upload resume first — Workday auto-parses it
    if safe_upload(page,
        "[data-automation-id='file-upload-input']",
        resume_path, timeout=5000):
        filled.append("✅ Resume")
        print("  ⏳ Waiting for Workday to parse resume...")
        time.sleep(5)
    elif safe_upload(page, "input[type='file']",
                     resume_path):
        filled.append("✅ Resume")
        time.sleep(4)

    # Basic info — Workday uses data-automation-id
    if try_fill_any(page, [
        "[data-automation-id='legalNameSection_firstName']",
        "[data-automation-id='firstName']",
        "input[placeholder*='First' i]"
    ], data.get("first_name", "")):
        filled.append("First name")

    if try_fill_any(page, [
        "[data-automation-id='legalNameSection_lastName']",
        "[data-automation-id='lastName']",
        "input[placeholder*='Last' i]"
    ], data.get("last_name", "")):
        filled.append("Last name")

    if try_fill_any(page, [
        "[data-automation-id='email']",
        "input[type='email']",
        "input[placeholder*='Email' i]"
    ], data.get("email", "")):
        filled.append("Email")

    if try_fill_any(page, [
        "[data-automation-id='phone']",
        "input[placeholder*='Phone' i]",
        "input[type='tel']"
    ], data.get("phone", "")):
        filled.append("Phone")

    if try_fill_any(page, [
        "[data-automation-id='city']",
        "input[placeholder*='City' i]"
    ], data.get("city", "")):
        filled.append("City")

    # Education
    edu_list = data.get("education", [])
    if edu_list:
        edu = edu_list[0]
        if try_fill_any(page, [
            "[data-automation-id='school']",
            "input[placeholder*='School' i]",
            "input[placeholder*='University' i]"
        ], edu.get("university", "")):
            filled.append("University")

        if try_fill_any(page, [
            "[data-automation-id='degree']",
            "input[placeholder*='Degree' i]"
        ], edu.get("degree", "")):
            filled.append("Degree")

        if try_fill_any(page, [
            "[data-automation-id='fieldOfStudy']",
            "input[placeholder*='Field' i]",
            "input[placeholder*='Major' i]"
        ], edu.get("field", "")):
            filled.append("Field")

    # Work experience
    exp_list = data.get("work_experience", [])
    if exp_list:
        exp = exp_list[0]
        if try_fill_any(page, [
            "[data-automation-id='employer']",
            "input[placeholder*='Employer' i]",
            "input[placeholder*='Company' i]"
        ], exp.get("company", "")):
            filled.append("Employer")

        if try_fill_any(page, [
            "[data-automation-id='title']",
            "input[placeholder*='Title' i]",
            "input[placeholder*='Position' i]"
        ], exp.get("title", "")):
            filled.append("Title")

    # EEO fields on Workday
    if try_select_any(page, [
        "select[data-automation-id*='gender' i]",
        "select[id*='gender' i]"
    ], candidate.get("eeo_gender", "Female")):
        filled.append("Gender")

    if try_select_any(page, [
        "select[data-automation-id*='ethnicity' i]",
        "select[data-automation-id*='race' i]",
        "select[id*='race' i]",
        "select[id*='ethnicity' i]"
    ], candidate.get("eeo_race", "Asian")):
        filled.append("Race")

    if try_select_any(page, [
        "select[data-automation-id*='veteran' i]",
        "select[id*='veteran' i]"
    ], candidate.get("eeo_veteran",
                     "I am not a protected veteran")):
        filled.append("Veteran")

    if try_select_any(page, [
        "select[data-automation-id*='disability' i]",
        "select[id*='disability' i]"
    ], candidate.get("eeo_disability",
                     "I don't wish to answer")):
        filled.append("Disability")

    print(f"  ✅ Workday filled: {', '.join(filled)}")
    return filled


# -------------------------------------------------------------------
# LEVER FILLER
# Confirmed selectors from Lever docs:
# input[name='name'], input[name='email'], input[name='phone']
# input[name='org'], input[name='urls[LinkedIn]']
# input[name='urls[GitHub]'], input[type='file']
# -------------------------------------------------------------------
def fill_lever(page, data, candidate, resume_path):
    print("  📋 Filling Lever form...")
    filled = []

    page.wait_for_load_state("domcontentloaded")
    time.sleep(1)

    # Full name (Lever uses single name field)
    if safe_fill(page, "input[name='name']",
                 data.get("full_name", "")):
        filled.append("Full name")

    if safe_fill(page, "input[name='email']",
                 data.get("email", "")):
        filled.append("Email")

    if safe_fill(page, "input[name='phone']",
                 data.get("phone", "")):
        filled.append("Phone")

    # Current company from work experience
    exp_list = data.get("work_experience", [])
    if exp_list:
        if safe_fill(page, "input[name='org']",
                     exp_list[0].get("company", "")):
            filled.append("Company")

    # LinkedIn and GitHub
    if safe_fill(page, "input[name='urls[LinkedIn]']",
                 data.get("linkedin_url", "")):
        filled.append("LinkedIn")

    if safe_fill(page, "input[name='urls[GitHub]']",
                 data.get("github_url", "")):
        filled.append("GitHub")

    # Resume upload
    if safe_upload(page, "input[type='file']", resume_path):
        filled.append("✅ Resume")
    elif safe_upload(page, "input[name='resume']",
                     resume_path):
        filled.append("✅ Resume")

    # EEO on Lever (if present)
    if try_select_any(page, [
        "select[id*='gender' i]",
        "select[name*='gender' i]"
    ], candidate.get("eeo_gender", "Female")):
        filled.append("Gender")

    if try_select_any(page, [
        "select[id*='race' i]",
        "select[id*='ethnicity' i]"
    ], candidate.get("eeo_race", "Asian")):
        filled.append("Race")

    if try_select_any(page, [
        "select[id*='veteran' i]"
    ], candidate.get("eeo_veteran",
                     "I am not a protected veteran")):
        filled.append("Veteran")

    if try_select_any(page, [
        "select[id*='disability' i]"
    ], candidate.get("eeo_disability",
                     "I don't wish to answer")):
        filled.append("Disability")

    print(f"  ✅ Lever filled: {', '.join(filled)}")
    return filled


# -------------------------------------------------------------------
# UNKNOWN PORTAL
# -------------------------------------------------------------------
def handle_unknown(page, data, candidate):
    print("\n  ℹ️  Unknown portal — use reference panel above.")
    print("  Fill fields manually then click Submit.")


# -------------------------------------------------------------------
# HIGHLIGHT FIELDS
# Yellow = unfilled required, Green = filled, Green button = Submit
# -------------------------------------------------------------------
def highlight_fields(page):
    page.evaluate("""
    () => {
        // Highlight all filled inputs green
        document.querySelectorAll('input, textarea').forEach(el => {
            if (el.value && el.value.trim() !== '') {
                el.style.backgroundColor = '#E8F5E9';
                el.style.border = '1px solid #4CAF50';
            }
        });

        // Highlight filled selects green
        document.querySelectorAll('select').forEach(el => {
            if (el.value && el.value !== '') {
                el.style.backgroundColor = '#E8F5E9';
                el.style.border = '1px solid #4CAF50';
            } else {
                el.style.backgroundColor = '#FFEB3B';
                el.style.border = '2px solid #F57F17';
            }
        });

        // Yellow for empty required inputs
        document.querySelectorAll(
            'input[required], textarea[required], ' +
            'input[aria-required="true"]'
        ).forEach(el => {
            if (!el.value || el.value.trim() === '') {
                el.style.backgroundColor = '#FFEB3B';
                el.style.border = '2px solid #F57F17';
            }
        });

        // Green Submit button
        const submitSels = [
            'button[type="submit"]',
            'input[type="submit"]',
            'button[data-submit]',
            '[data-automation-id="bottom-navigation-next-button"]'
        ];
        submitSels.forEach(sel => {
            const btn = document.querySelector(sel);
            if (btn) {
                btn.style.border = '4px solid #00AA00';
                btn.style.backgroundColor = '#90EE90';
                btn.style.color = '#000';
                btn.style.fontWeight = 'bold';
                btn.scrollIntoView({behavior: 'smooth'});
            }
        });
    }
    """)
    print("  🟡 Yellow = fill these  |  🟢 Green = filled  "
          "|  🟢 Button = Submit")


# -------------------------------------------------------------------
# UPDATE EXCEL
# -------------------------------------------------------------------
def update_applied_status(job):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Job Applications"]
        title   = job.get("title", "")
        company = job.get("company_name",
                          job.get("company", ""))
        for row in ws.iter_rows(min_row=2,
                                max_row=ws.max_row):
            if (row[2].value == title and
                    row[3].value == company):
                row[8].value  = "Yes"
                row[9].value  = datetime.now().strftime(
                    "%Y-%m-%d"
                )
                row[10].value = "Applied"
                break
        wb.save(EXCEL_FILE)
        print("  📊 Excel updated — Applied: Yes")
    except Exception as e:
        print(f"  ⚠️  Excel error: {e}")


# -------------------------------------------------------------------
# MAIN APPLY FUNCTION
# -------------------------------------------------------------------
def apply_to_job(result):
    job         = result.get("job", result)
    resume_path = result.get("file", "")
    data        = get_resume_data(result)
    candidate   = load_candidate()

    # Get apply link
    apply_link = result.get("apply", "")
    if not apply_link:
        opts = job.get("apply_options", [])
        if opts:
            apply_link = opts[0].get("link", "")

    if not apply_link:
        print("  ⚠️  No apply link found")
        return False

    if not data:
        print("  ⚠️  No app_data found — run main.py first")
        return False

    if not resume_path or not os.path.exists(resume_path):
        print(f"  ⚠️  Resume not found: {resume_path}")
        return False

    portal = detect_portal(apply_link)
    title  = job.get('title', result.get('title', ''))
    co     = job.get('company_name',
                     result.get('company', ''))

    print(f"\n  {'='*55}")
    print(f"  🌐 {title} @ {co}")
    print(f"  Portal:  {portal.upper()}")
    print(f"  Resume:  {os.path.basename(resume_path)}")
    print(f"  ⏳ ~15-20 seconds to fill...")
    print(f"  {'='*55}")

    # Show reference panel
    show_reference_panel(data, candidate)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page    = context.new_page()

        try:
            page.goto(apply_link, timeout=30000)
            page.wait_for_load_state("domcontentloaded")
            time.sleep(1.5)

            # Fill based on portal
            if portal == "greenhouse":
                fill_greenhouse(
                    page, data, candidate, resume_path
                )
            elif portal == "workday":
                fill_workday(
                    page, data, candidate, resume_path
                )
            elif portal == "lever":
                fill_lever(
                    page, data, candidate, resume_path
                )
            else:
                handle_unknown(page, data, candidate)

            # Highlight all fields
            time.sleep(0.5)
            highlight_fields(page)

            print(f"\n  ✋ YOUR TURN:")
            print(f"  🟡 YELLOW = fill these manually")
            print(f"  🟢 GREEN  = already filled by agent")
            print(f"  🟢 GREEN button = click to Submit\n")

            input("  Press Enter AFTER you submit: ")

            update_applied_status(job)
            print("  ✅ Done and logged!")
            return True

        except Exception as e:
            print(f"  ❌ Error: {e}")
            print(f"  ℹ️  Apply manually: {apply_link}")
            input("  Press Enter when done: ")
            return False

        finally:
            browser.close()


# -------------------------------------------------------------------
# BATCH APPLY
# -------------------------------------------------------------------
def run_auto_apply(results):
    if not results:
        print("  ⚠️  No jobs to apply to!")
        return

    print(f"\n{'='*55}")
    print(f"  🚀 PHASE 3: Semi-Auto Apply")
    print(f"  {len(results)} jobs ready")
    print(f"{'='*55}\n")

    applied = 0
    skipped = 0

    for i, result in enumerate(results):
        title   = result.get("title", "")
        company = result.get("company", "")
        score   = result.get("score",
                  result.get("total_score", 0))
        file    = result.get("file", "")

        print(f"\n  [{i+1}/{len(results)}] {title} @ {company}")
        print(f"  Score: {score}/100")
        print(f"  Apply: {result.get('apply', '')[:55]}")

        choice = input(
            "\n  Apply? (y/n/q to quit): "
        ).strip().lower()

        if choice == 'q':
            break
        elif choice == 'n':
            skipped += 1
        elif choice == 'y':
            if apply_to_job(result):
                applied += 1
            else:
                skipped += 1

    print(f"\n{'='*55}")
    print(f"  ✅ Session done!")
    print(f"  Applied: {applied} | Skipped: {skipped}")
    print(f"{'='*55}\n")


# -------------------------------------------------------------------
# TEST
# -------------------------------------------------------------------
if __name__ == "__main__":
    # Find real resume
    resume_path = ""
    if os.path.exists("tailored_resumes"):
        resumes = sorted([
            f for f in os.listdir("tailored_resumes")
            if f.endswith(".pdf")
        ])
        if resumes:
            resume_path = f"tailored_resumes/{resumes[0]}"

    if not resume_path:
        print("⚠️  No resumes found. Run python main.py first.")
        exit()

    print(f"Testing with: {resume_path}\n")

    # Load real profile
    candidate = load_candidate()

    test_result = {
        "title":   "Business Analyst",
        "company": "Test Company",
        "score":   88,
        "type":    "FULLTIME",
        "file":    resume_path,
        "apply":   "https://job-boards.greenhouse.io/anthropic/jobs/5127289008",
        "job": {
            "title":        "Business Analyst",
            "company_name": "Test Company",
            "apply_options": [{
                "link": "https://job-boards.greenhouse.io/anthropic/jobs/5127289008"
            }]
        },
        "app_data": {
            "first_name":   candidate.get("first_name", "Hitankshi"),
            "last_name":    candidate.get("last_name", "Jain"),
            "full_name":    candidate.get("full_name", "Hitankshi Jain"),
            "email":        candidate.get("email",
                            "Jain.hitankshi09@gmail.com"),
            "phone":        candidate.get("phone", "445-260-9057"),
            "location":     candidate.get("location",
                            "Philadelphia, PA"),
            "city":         candidate.get("city", "Philadelphia"),
            "state":        candidate.get("state", "PA"),
            "linkedin_url": candidate.get("linkedin_url",
                            "https://www.linkedin.com/in/hitankshi-jain"),
            "github_url":   candidate.get("github_url",
                            "https://github.com/hitankshi09"),
            "education": [{
                "university": "Drexel University",
                "degree":     "MS",
                "field":      "Business Analytics",
                "start_date": "September 2023",
                "end_date":   "March 2025",
                "location":   "Philadelphia, USA",
                "gpa":        ""
            }],
            "work_experience": [{
                "company":    "Neuralix.ai",
                "title":      "Business Analyst",
                "start_date": "July 2025",
                "end_date":   "Present",
                "location":   "Philadelphia, USA",
                "bullets":    []
            }],
            "skills": [
                "Python", "SQL", "PowerBI",
                "Tableau", "R", "Excel"
            ],
            "tailored_bullets": [
                "Developed KPI dashboards in PowerBI and Tableau "
                "monitoring 50K+ records",
                "Collaborated with cross-functional teams using agile",
                "Automated data pipelines using SQL and Python APIs"
            ]
        }
    }

    apply_to_job(test_result)
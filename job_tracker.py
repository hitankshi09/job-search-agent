import json
import os
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from datetime import datetime, timedelta

TRACKER_FILE  = "applied_jobs.json"
EXCEL_FILE    = "job_search_tracker.xlsx"


# -------------------------------------------------------------------
# JSON TRACKER — remembers seen jobs across runs
# -------------------------------------------------------------------
def load_tracker():
    if os.path.exists(TRACKER_FILE):
        with open(TRACKER_FILE, "r") as f:
            return json.load(f)
    return {}


def save_tracker(tracker):
    with open(TRACKER_FILE, "w") as f:
        json.dump(tracker, f, indent=2)


def get_job_key(job):
    job_id = job.get("job_id", "")
    if job_id:
        return job_id
    title   = (job.get("title", "") or "").lower().replace(" ", "")
    company = (job.get("company_name", "") or "").lower().replace(" ", "")
    return f"{company}_{title}"


def is_already_seen(job):
    tracker = load_tracker()
    return get_job_key(job) in tracker


def filter_seen_jobs(jobs):
    tracker    = load_tracker()
    fresh_jobs = []
    skipped    = 0
    for job in jobs:
        if get_job_key(job) in tracker:
            skipped += 1
        else:
            fresh_jobs.append(job)
    if skipped > 0:
        print(f"  🔁 Skipped {skipped} previously seen jobs")
    return fresh_jobs


def mark_as_seen(job, score=None, verdict=None, resume_file=None, status="Resume Ready"):
    tracker = load_tracker()
    job_id  = get_job_key(job)

    apply_options = job.get("apply_options", [{}])
    apply_link    = apply_options[0].get("link", "") if apply_options else ""

    # Calculate follow-up dates
    today         = datetime.now()
    followup_date = (today + timedelta(days=2)).strftime("%Y-%m-%d")
    second_followup = (today + timedelta(days=7)).strftime("%Y-%m-%d")

    tracker[job_id] = {
        "title":            job.get("title", ""),
        "company":          job.get("company_name", ""),
        "location":         job.get("job_city", "") or job.get("location", ""),
        "job_type":         job.get("job_type", "").upper(),
        "posted_at":        job.get("posted_at", "")[:10],
        "first_seen":       today.strftime("%Y-%m-%d %H:%M"),
        "score":            score or 0,
        "verdict":          verdict or "",
        "apply_link":       apply_link,
        "resume_file":      resume_file or "",
        "applied":          "No",
        "applied_date":     "",
        "status":           status,
        "linkedin_followup_date":   followup_date,
        "linkedin_2nd_followup":    second_followup,
        "notes":            ""
    }
    save_tracker(tracker)
    return tracker[job_id]


def update_application_status(job_key, applied=True, status=None, notes=None):
    """Call this manually to mark a job as applied"""
    tracker = load_tracker()
    if job_key in tracker:
        if applied:
            tracker[job_key]["applied"] = "Yes"
            tracker[job_key]["applied_date"] = datetime.now().strftime("%Y-%m-%d")
        if status:
            tracker[job_key]["status"] = status
        if notes:
            tracker[job_key]["notes"] = notes
        save_tracker(tracker)
        update_excel()


# -------------------------------------------------------------------
# EXCEL TRACKER — beautiful spreadsheet updated every run
# -------------------------------------------------------------------
def setup_excel():
    """Create Excel file with headers if it doesn't exist"""
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Applications"

    # Column headers
    headers = [
        "No.",
        "Date Seen",
        "Job Title",
        "Company",
        "Location",
        "Type",
        "ATS Score",
        "Verdict",
        "Applied?",
        "Applied Date",
        "Status",
        "Apply Link",
        "Resume File",
        "LinkedIn Follow-up Date",    
        "LinkedIn 2nd Follow-up",     
        "Connection Sent?",           
        "Follow-up DM Sent?",         
        "Notes"
    ]

    # Header styling
    header_fill   = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    center_align  = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col, header in enumerate(headers, 1):
        cell              = ws.cell(row=1, column=col, value=header)
        cell.fill         = header_fill
        cell.font         = header_font
        cell.alignment    = center_align
        cell.border       = thin_border

    # Freeze top row so headers always visible
    ws.freeze_panes = "A2"

    # Set column widths
    col_widths = {
        "A": 5,   # No.
        "B": 12,  # Date Seen
        "C": 30,  # Job Title
        "D": 25,  # Company
        "E": 20,  # Location
        "F": 12,  # Type
        "G": 10,  # ATS Score
        "H": 16,  # Verdict
        "I": 10,  # Applied?
        "J": 14,  # Applied Date
        "K": 20,  # Status
        "L": 45,  # Apply Link
        "M": 45,  # Resume File
        "N": 30,  # Notes
        "O": 22,  # LinkedIn Follow-up Date
        "P": 22,  # LinkedIn 2nd Follow-up
        "Q": 16,  # Connection Sent?
        "R": 16,  # Follow-up DM Sent?
        "S": 30,  # Notes
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(EXCEL_FILE)
    return wb


def check_followups():
    """
    Checks who needs a LinkedIn follow-up today.
    Call this at the start of each day.
    """
    tracker = load_tracker()
    today   = datetime.now().strftime("%Y-%m-%d")

    first_followups  = []
    second_followups = []

    for job_id, job in tracker.items():
        # Check first follow-up (2 days after connection)
        if job.get("linkedin_followup_date") == today:
            first_followups.append(job)

        # Check second follow-up (7 days after connection)
        if job.get("linkedin_2nd_followup") == today:
            second_followups.append(job)

    if first_followups:
        print(f"\n  🔔 LINKEDIN FOLLOW-UP REMINDERS — Today {today}")
        print(f"  {'='*50}")
        print("\n  📩 First follow-up (send DM to those who accepted):\n")
        for job in first_followups:
            print(f"  • {job.get('title')} @ {job.get('company')}")
            print(f"    Connected on: {job.get('first_seen', '')[:10]}")
            print("    Check Excel for their follow-up DM\n")

    if second_followups:
        print("\n  🔔 Second follow-up (no response yet — try again):\n")
        for job in second_followups:
            print(f"  • {job.get('title')} @ {job.get('company')}")
            print(f"    Original connection: {job.get('first_seen', '')[:10]}")
            print("    Consider a gentle nudge!\n")

    if not first_followups and not second_followups:
        print("\n  ✅ No LinkedIn follow-ups due today!")

    return first_followups, second_followups

def update_excel():
    """Rebuild Excel from tracker JSON — called after every run"""
    tracker = load_tracker()

    # Load or create workbook
    wb = setup_excel()
    ws = wb["Job Applications"]

    # Clear existing data rows (keep header)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Color definitions
    strong_fill   = PatternFill("solid", fgColor="C6EFCE")  # green
    good_fill     = PatternFill("solid", fgColor="FFEB9C")  # yellow
    applied_fill  = PatternFill("solid", fgColor="BDD7EE")  # blue
    default_fill  = PatternFill("solid", fgColor="FFFFFF")  # white
    alt_fill      = PatternFill("solid", fgColor="F2F2F2")  # light grey

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    link_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Sort by score descending
    sorted_jobs = sorted(
        tracker.values(),
        key=lambda x: x.get("score", 0),
        reverse=True
    )

    for row_idx, job in enumerate(sorted_jobs, start=2):
        score   = job.get("score", 0)
        applied = job.get("applied", "No")

        # Row background color
        if applied == "Yes":
            row_fill = applied_fill
        elif score >= 80:
            row_fill = strong_fill
        elif score >= 70:
            row_fill = good_fill
        elif row_idx % 2 == 0:
            row_fill = alt_fill
        else:
            row_fill = default_fill

        # Row data
        row_data = [
            row_idx - 1,
            job.get("first_seen", "")[:10],
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("job_type", ""),
            score,
            job.get("verdict", ""),
            applied,
            job.get("applied_date", ""),
            job.get("status", ""),
            job.get("apply_link", ""),
            job.get("resume_file", ""),
            job.get("linkedin_followup_date", ""),   # ← new
            job.get("linkedin_2nd_followup", ""),    # ← new
            job.get("connection_sent", "No"),        # ← new
            job.get("followup_sent", "No"),          # ← new
            job.get("notes", "")
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = thin_border

            # Alignment per column
            if col_idx in [1, 2, 6, 7, 8, 9, 10]:
                cell.alignment = center
            elif col_idx in [12, 13]:
                cell.alignment = link_align
            else:
                cell.alignment = left

            # Bold job title
            if col_idx == 3:
                cell.font = Font(bold=True, size=10)

            # Color score cell
            if col_idx == 7:
                if score >= 80:
                    cell.font = Font(bold=True, color="375623", size=10)
                elif score >= 70:
                    cell.font = Font(bold=True, color="7D6608", size=10)
                else:
                    cell.font = Font(size=10)

            # Color applied cell
            if col_idx == 9:
                if applied == "Yes":
                    cell.font = Font(bold=True, color="1F4E79", size=10)
                else:
                    cell.font = Font(color="C00000", size=10)

    # Add summary row at bottom
    summary_row = ws.max_row + 2
    total       = len(sorted_jobs)
    applied_cnt = len([j for j in sorted_jobs if j.get("applied") == "Yes"])
    strong_cnt  = len([j for j in sorted_jobs if j.get("score", 0) >= 80])
    pending_cnt = len([j for j in sorted_jobs
                       if j.get("applied") == "No"
                       and j.get("score", 0) >= 80])

    summary_fill = PatternFill("solid", fgColor="1F4E79")
    summary_font = Font(color="FFFFFF", bold=True, size=10)

    summaries = [
        (1,  f"Total: {total}"),
        (3,  f"Strong (≥80): {strong_cnt}"),
        (5,  f"Applied: {applied_cnt}"),
        (7,  f"Pending strong: {pending_cnt}"),
        (9,  f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"),
    ]

    for col, text in summaries:
        cell           = ws.cell(row=summary_row, column=col, value=text)
        cell.fill      = summary_fill
        cell.font      = summary_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(EXCEL_FILE)
    print(f"  📊 Excel tracker updated: {EXCEL_FILE}")


def show_tracker_summary():
    tracker = load_tracker()
    total    = len(tracker)
    applied  = len([j for j in tracker.values() if j.get("applied") == "Yes"])
    strong   = len([j for j in tracker.values() if j.get("score", 0) >= 80])
    pending  = len([j for j in tracker.values()
                    if j.get("applied") == "No" and j.get("score", 0) >= 80])

    print("\n  📋 Job Tracker Summary:")
    print(f"     Total jobs tracked:    {total}")
    print(f"     Strong matches (≥80):  {strong}")
    print(f"     Applied so far:        {applied}")
    print(f"     Strong but pending:    {pending}")
    print(f"     📊 Full tracker:       {EXCEL_FILE}")